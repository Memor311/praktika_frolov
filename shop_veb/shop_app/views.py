from django.db.models import Q, Sum, Avg, F, Count
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from .forms import CustomUserRegistrationForm, ProductForm, SiteReviewForm, UserProfileForm, ReviewForm, CustomAuthenticationForm
from shop_app.models import Product, Category, Order, Review, Status, ReviewPhoto, OrderProduct
from django.contrib.auth.views import LoginView
from django.urls import reverse_lazy
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from django.contrib.auth.decorators import user_passes_test
from datetime import datetime

def admin_required(view_func):
    def check_admin(user):
        return user.is_authenticated and hasattr(user, 'role') and user.role.name == 'Администратор'
    
    return user_passes_test(check_admin, login_url='catalog')(view_func)



def index(request):
    reviews = Review.objects.all()
    return render(request, 'index.html', {'reviews': reviews})

@login_required
def add_site_review_view(request):
    if request.method == 'POST':
        form = SiteReviewForm(request.POST, request.FILES)
        if form.is_valid():
            # Сохраняем отзыв
            review = form.save(commit=False)
            review.user = request.user
            review.save()
            photos = request.FILES.getlist('photos')
            for photo in photos:
                ReviewPhoto.objects.create(review=review, photo=photo)
            messages.success(request, "Спасибо за ваш отзыв!")
            return redirect('index')
        else:
            messages.error(request, "Пожалуйста, исправьте ошибки.")
    else:
        form = SiteReviewForm()
    
    return render(request, 'add_site_review.html', {'form': form})


def catalog_view(request):
    products = Product.objects.select_related('category').filter(is_active=True)
    category = Category.objects.all()
    search = request.GET.get('search','')
    if search:
        products=products.filter(name__icontains=search)
    filtered=request.GET.get('filter','')
    if filtered:
        products = products.filter(category_id=filtered)
    
    context ={
        'products': products,
        'category': category,
        'search': search,
        'filter': filtered,
    }
    return render(request, 'products/catalog.html', context)


def product_detail_view(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    return render(request, 'products/product_detail.html', {'product': product})

class CustomLoginView(LoginView):
    form_class = CustomAuthenticationForm
    template_name = 'login/login.html'
    redirect_authenticated_user = True

    def get_success_url(self):
        return reverse_lazy('catalog')

def register_view(request):
    if request.method == 'POST':
        form = CustomUserRegistrationForm(request.POST)
        if form.is_valid():
            form.save()
            username = form.cleaned_data.get('username')
            messages.success(request, f'Аккаунт {username} успешно создан! Теперь вы можете войти.')
            return redirect('login')
        else:
            messages.error(request, 'Пожалуйста, исправьте ошибки в форме.')
    else:
        form = CustomUserRegistrationForm()
    return render(request, 'login/register.html', {'form': form})


@admin_required
def admin_dashboard_view(request):
    return render(request, 'admin-panel/admin_dashboard.html')

@admin_required
def admin_products_view(request):
    products = Product.objects.select_related('category').all()
    return render(request, 'admin-panel/products/admin_products.html', {'products': products})

@admin_required
def admin_orders_view(request):
    orders = Order.objects.select_related('user', 'status').prefetch_related('order_items__product')
    search_query = request.GET.get('search', '')
    if search_query:
        orders = orders.filter(
            Q(user__username__icontains=search_query) |
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query)
        )

    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    if date_from:
        orders = orders.filter(created_at__gte=date_from)
    if date_to:
        orders = orders.filter(created_at__lte=date_to + ' 23:59:59')
    return render(request, 'admin-panel/orders/admin_orders.html', {
        'orders': orders,
        'statuses': Status.objects.all(),
        'search_query': search_query,
        'date_from': date_from,
        'date_to': date_to,
    })

@admin_required
def admin_reviews_view(request):
    reviews = Review.objects.select_related('user').prefetch_related('photos').order_by('-created_at')
    return render(request, 'admin-panel/reviews/admin_reviews.html', {'reviews': reviews})

@admin_required 
def add_product_view(request):
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            product = form.save(commit=False)
            product.creator = request.user
            product.save()
            messages.success(request, f'Товар «{product.name}» успешно добавлен!')
            return redirect('admin-panel')
        else:
            messages.error(request, 'Пожалуйста, исправьте ошибки.')
    else:
        form = ProductForm()
    return render(request, 'admin-panel/products/add_product.html', {'form': form})


@admin_required
def edit_product_view(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            form.save()
            messages.success(request, f'Товар «{product.name}» обновлён!')
            return redirect('admin-panel')
        else:
            messages.error(request, 'Пожалуйста, исправьте ошибки.')
    else:
        form = ProductForm(instance=product)
    return render(request, 'admin-panel/products/edit_product.html', {'form': form, 'product': product})

@admin_required
def delete_product_view(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    if product.is_active:
        product.is_active = False
        product.save()
        messages.success(request, f'Товар скрыт из каталога.')
    else:
        product.is_active = True
        product.save()
        messages.success(request, f'Товар вернут в каталог.')
    return redirect('admin-products')

@admin_required
def create_order_view(request):
    if request.method == 'POST':
        address = request.POST.get('address')
        product_ids = request.POST.getlist('product_ids')
        counts = request.POST.getlist('counts')            
        order = Order.objects.create(
            user=request.user,
            address=address,
            status_id=1
        )
        for prod_id, count in zip(product_ids, counts):
            product = Product.objects.get(id=prod_id)
            OrderProduct.objects.create(
                order=order,
                product=product,
                count=int(count)
            )
        order.calculate_total()
        return redirect('order-success', order_id=order.id)
    cart = request.session.get('cart', {})
    products = Product.objects.filter(id__in=cart.keys())
    return render(request, 'orders/create_order.html', {'products': products, 'cart': cart})

def order_detail_view(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    if not request.user.is_admin and order.user != request.user:
        messages.error(request, "У вас нет доступа к этому заказу.")
        return redirect('catalog')
    return render(request, 'orders/order_detail.html', {'order': order})

@login_required
def delete_review_view(request, review_id):
    review = Review.objects.get(id=review_id)
    review.delete()
    return redirect('admin-reviews')

def add_to_cart(request):
    if request.method == 'POST':
        product_id = request.POST.get('product_id')
        product = get_object_or_404(Product, id=product_id)
        cart = request.session.get('cart', {})
        cart[product_id] = cart.get(product_id, 0) + 1
        request.session['cart'] = cart
        request.session.modified = True
        messages.success(request, f'«{product.name}» добавлен в корзину!')

    referer = request.META.get('HTTP_REFERER')
    if referer:
        return redirect(referer)
    else:
        # Если нет Referer — резервный вариант
        return redirect('catalog')

def cart_view(request):
    cart = request.session.get('cart', {})
    if request.method == 'POST':
        product_id = request.POST.get('product_id')
        action = request.POST.get('action')
        if product_id in cart:
            if action == 'increase':
                cart[product_id] += 1
            elif action == 'decrease':
                if cart[product_id] > 1:
                    cart[product_id] -= 1
                else:
                    del cart[product_id]
            elif action == 'remove':
                del cart[product_id]
            request.session['cart'] = cart
            request.session.modified = True
            return redirect('cart')
    cart_items = []
    total = 0
    for product_id, count in cart.items():
        try:
            product = Product.objects.get(id=product_id)
            item_total = product.price * count
            cart_items.append({
                'product': product,
                'count': count,
                'total': item_total
            })
            total += item_total
        except Product.DoesNotExist:
            continue
    return render(request, 'cart.html', {
        'cart_items': cart_items,
        'total': total
    })

def checkout_view(request):
    if not request.user.is_authenticated:
        messages.error(request, "Для оформления заказа нужно войти в аккаунт.")
        return redirect('login')
    cart = request.session.get('cart', {})
    if not cart:
        messages.error(request, "Корзина пуста.")
        return redirect('cart')
    initial_address = ""
    last_order = request.user.orders.order_by('-created_at').first()
    if last_order and last_order.address:
        initial_address = last_order.address
    if request.method == 'POST':
        address = request.POST.get('address', '').strip()
        if not address:
            messages.error(request, "Укажите адрес доставки.")
            return render(request, 'checkout.html', {'initial_address': initial_address})
        status_new = Status.objects.filter(name='Новый').first()
        if not status_new:
            status_new = Status.objects.create(name='Новый')
        order = Order.objects.create(
            user=request.user,
            address=address,
            status=status_new
        )
        for product_id, count in cart.items():
            try:
                product = Product.objects.get(id=product_id)
                OrderProduct.objects.create(order=order, product=product, count=int(count))
            except Product.DoesNotExist:
                continue
        order.calculate_total()
        request.session['cart'] = {}
        messages.success(request, f"Заказ №{order.id} успешно оформлен!")
        return redirect('order-detail', order_id=order.id)
    return render(request, 'checkout.html', {'initial_address': initial_address})

@login_required
def profile_view(request):
    if request.method == 'POST':
        form = UserProfileForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, "Данные профиля обновлены!")
            return redirect('profile')
    else:
        form = UserProfileForm(instance=request.user)
    orders = request.user.orders.select_related('status').all().order_by('-created_at')
    return render(request, 'login/profile.html', {
        'form': form,
        'orders': orders
    })


def update_order_status_view(request, order_id):
    if request.method != 'POST':
        messages.error(request, "Недопустимый метод запроса.")
        return redirect('admin-panel')
    order = get_object_or_404(Order, id=order_id)
    status_id = request.POST.get('status_id')
    if not status_id:
        messages.error(request, "Статус не выбран.")
        return redirect('admin-orders')
    try:
        new_status = Status.objects.get(id=status_id)
        order.status = new_status
        order.save(update_fields=['status'])
        messages.success(request, f"Статус заказа #{order.id} обновлён на «{new_status.name}».")
    except Status.DoesNotExist:
        messages.error(request, "Указанный статус не существует.")
    return redirect('admin-orders')

def sales_report_view(request):
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    orders_qs = Order.objects.all()
    if date_from:
        orders_qs = orders_qs.filter(created_at__date__gte=date_from)
    if date_to:
        orders_qs = orders_qs.filter(created_at__date__lte=date_to)

    report_data = (
        OrderProduct.objects
        .filter(order__in=orders_qs)
        .select_related('product', 'product__category')
        .values('product__id', 'product__name', 'product__category__name')
        .annotate(
            total_quantity=Sum('count'),
            total_revenue=Sum(F('count') * F('product__price')),
            avg_per_order=Avg('count'),
            order_count=Count('order', distinct=True),
        )
        .order_by('-total_revenue')
    )
    report = []
    for item in report_data:
        report.append({
            'product_name': item['product__name'],
            'category_name': item['product__category__name'],
            'total_quantity': item['total_quantity'] or 0,
            'total_revenue': float(item['total_revenue'] or 0),
            'avg_per_order': float(item['avg_per_order'] or 0),
            'order_count': item['order_count'],
        })

    if request.GET.get('export') == 'xlsx':
        return export_sales_report_xlsx(report)
    context = {
        'report': report,
        'date_from': date_from,
        'date_to': date_to,
    }
    return render(request, 'admin-panel/sales_report.html', context)


def export_sales_report_xlsx(report, date_from=None, date_to=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт по продажам"
    def fmt_date(d):
        if not d:
            return None
        try:
            dt = datetime.strptime(d, '%Y-%m-%d')
            return dt.strftime('%d.%m.%Y')
        except ValueError:
            return d
    formatted_from = fmt_date(date_from)
    formatted_to = fmt_date(date_to)
    if formatted_from and formatted_to:
        title = f"Отчёт с {formatted_from} по {formatted_to}"
    elif formatted_from:
        title = f"Отчёт с {formatted_from}"
    elif formatted_to:
        title = f"Отчёт по {formatted_to}"
    else:
        title = "Отчёт за всё время"
    ws.merge_cells('A1:D1')
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")
    headers = ['Товар', 'Количество покупок', 'Сумма (₽)', 'Среднее кол-во в заказе']
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for item in report:
        ws.append([
            item.get('product_name', ''),
            item.get('total_quantity', 0),
            item.get('total_revenue', 0),
            item.get('avg_per_order', 0),
        ])
    col_widths = [0, 0, 0, 0]  # для A, B, C, D
    for i, h in enumerate(headers):
        col_widths[i] = max(col_widths[i], len(h))
    for row in ws.iter_rows(min_row=3):
        for i, cell in enumerate(row):
            if i >= 4:
                break
            if cell.value is not None:
                try:
                    length = len(str(cell.value))
                    col_widths[i] = max(col_widths[i], length)
                except:
                    pass
    for i, width in enumerate(col_widths):
        letter = get_column_letter(i + 1)  # A=1, B=2, ...
        ws.column_dimensions[letter].width = min(width + 2, 50)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=sales_report.xlsx'
    wb.save(response)
    return response

def admin_order_detail(request, order_id):
    order = get_object_or_404(Order.objects.select_related('user'), id=order_id)
    return render(request, 'admin-panel/orders/order_detail.html', {'order': order})    

@admin_required
def edit_order_view(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add_product':
            product_id = request.POST.get('new_product_id')
            count = request.POST.get('count', '1')
            try:
                count = int(count)
                if count < 1:
                    raise ValueError
                product = Product.objects.get(id=product_id)
                order_item, created = OrderProduct.objects.get_or_create(
                    order=order,
                    product=product,
                    defaults={'count': count}
                )
                if not created:
                    order_item.count += count
                    order_item.save()
                messages.success(request, f"Товар «{product.name}» добавлен ({count} шт.).")
                order.update_total_price()
                return redirect('edit-order', order_id=order.id)
            except (ValueError, Product.DoesNotExist):
                messages.error(request, "Некорректное количество или товар не найден.")
                return redirect('edit-order', order_id=order.id)
        product_id = request.POST.get('product_id')
        if action in ['increase', 'decrease', 'remove'] and product_id:
            order_item = get_object_or_404(OrderProduct, order=order, product_id=product_id)
            if action == 'increase':
                order_item.count += 1
                order_item.save()
                messages.success(request, f"Количество «{order_item.product.name}» увеличено.")
            elif action == 'decrease':
                if order_item.count > 1:
                    order_item.count -= 1
                    order_item.save()
                    messages.success(request, f"Количество «{order_item.product.name}» уменьшено.")
                else:
                    order_item.delete()
                    messages.info(request, f"Товар «{order_item.product.name}» удалён из заказа.")
            elif action == 'remove':
                product_name = order_item.product.name
                order_item.delete()
                messages.info(request, f"Товар «{product_name}» удалён из заказа.")
            order.update_total_price()
            return redirect('edit-order', order_id=order.id)
    all_products = Product.objects.filter(is_active=True).order_by('name')
    order_items = []
    for item in order.order_items.all():
        order_items.append({
            'product': item.product,
            'count': item.count,
            'total': item.total_price
        })
    total = sum(item['total'] for item in order_items)
    return render(request, 'admin-panel/orders/edit_order.html', {
        'order': order,
        'order_items': order_items,
        'total': total,
        'all_products': all_products,
    })

@admin_required
def delete_order_view(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    order.delete()
    return redirect('admin-orders')
